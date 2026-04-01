#!/usr/bin/env python3
"""Generate a policy-analysis Excel workbook from canonical summaries."""

from __future__ import annotations

import json
import os
from datetime import datetime
from typing import Any, Dict, List, Mapping

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from summary_schema import (
    canonical_ai_summary_json,
    get_core_analysis,
    get_primary_sheet_target,
    get_quality_audit,
    get_routing,
)


SUMMARY_FILE = "output/政策分析/政策分析_summaries.json"
OUTPUT_FILE = "output/政策分析/政策分析_analyzed_papers.xlsx"


def _stringify_list(value: Any) -> str:
    if isinstance(value, list):
        return "; ".join(str(item).strip() for item in value if str(item).strip())
    if value is None:
        return ""
    return str(value).strip()


def _authors_to_text(paper_info: Mapping[str, Any]) -> str:
    authors = paper_info.get("authors", [])
    if isinstance(authors, list):
        return ", ".join(str(author).strip() for author in authors if str(author).strip())
    return str(authors or "").strip()


def _build_record(summary: Mapping[str, Any]) -> Dict[str, Any]:
    paper_info = summary.get("paper_info", {}) if isinstance(summary, Mapping) else {}
    preprocess = summary.get("preprocess", {}) if isinstance(summary, Mapping) else {}
    routing = get_routing(summary)
    core = get_core_analysis(summary)
    quality = get_quality_audit(summary)

    return {
        "论文标题": paper_info.get("title", ""),
        "作者": _authors_to_text(paper_info),
        "发表年份": paper_info.get("year", ""),
        "期刊名称": paper_info.get("journal", ""),
        "DOI": paper_info.get("doi", ""),
        "文本长度": summary.get("text_length", 0),
        "主类型": routing.get("paper_type") or "",
        "分类状态": routing.get("classification_status") or "",
        "主类型工作表": get_primary_sheet_target(summary),
        "研究摘要": core.get("summary") or "",
        "关键要点": _stringify_list(core.get("key_points", [])),
        "研究方法": core.get("methodology") or "",
        "主要发现": core.get("findings") or "",
        "研究结论": core.get("conclusions") or "",
        "理论贡献": core.get("relevance") or "",
        "研究局限": core.get("limitations") or "",
        "提取置信度": quality.get("extraction_confidence") or "",
        "完整度": quality.get("completeness_score") or 0,
        "建议人工复核": bool(quality.get("needs_manual_review")),
        "处理状态": summary.get("status", ""),
        "处理时间": summary.get("processing_time", ""),
        "解析器": preprocess.get("extractor_used", "") or summary.get("engine_used", ""),
        "详细信息": canonical_ai_summary_json(summary),
    }


def _build_stats_rows(summaries: List[Mapping[str, Any]]) -> List[Dict[str, Any]]:
    total = len(summaries)
    success = sum(1 for item in summaries if item.get("status") == "success")
    failed = sum(1 for item in summaries if item.get("status") == "failed")
    manual_review = sum(1 for item in summaries if get_quality_audit(item).get("needs_manual_review"))
    empirical = sum(1 for item in summaries if get_routing(item).get("paper_type") == "empirical")
    review = sum(1 for item in summaries if get_routing(item).get("paper_type") == "review")
    conceptual = sum(1 for item in summaries if get_routing(item).get("paper_type") == "conceptual")

    return [
        {"统计项目": "总论文数", "数值": total},
        {"统计项目": "成功处理", "数值": success},
        {"统计项目": "失败处理", "数值": failed},
        {"统计项目": "成功率(%)", "数值": f"{(success / total * 100):.1f}%" if total else "0.0%"},
        {"统计项目": "建议人工复核", "数值": manual_review},
        {"统计项目": "实证论文", "数值": empirical},
        {"统计项目": "综述论文", "数值": review},
        {"统计项目": "概念论文", "数值": conceptual},
        {"统计项目": "项目名称", "数值": "政策分析"},
        {"统计项目": "生成时间", "数值": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
    ]


def create_excel_for_policy_analysis() -> bool:
    """Create the project-specific Excel workbook."""
    print("正在为政策分析项目生成 Excel 文件...")

    if not os.path.exists(SUMMARY_FILE):
        print(f"未找到摘要文件: {SUMMARY_FILE}")
        return False

    with open(SUMMARY_FILE, "r", encoding="utf-8") as handle:
        summaries = json.load(handle)

    if not summaries:
        print("没有找到可导出的摘要数据")
        return False

    print(f"找到 {len(summaries)} 篇论文")
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)

    records = [_build_record(summary) for summary in summaries]
    stats_rows = _build_stats_rows(summaries)
    df = pd.DataFrame(records)
    stats_df = pd.DataFrame(stats_rows)

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="论文分析摘要", index=False)
        stats_df.to_excel(writer, sheet_name="项目统计", index=False)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        summary_sheet = writer.sheets["论文分析摘要"]
        stats_sheet = writer.sheets["项目统计"]

        column_widths = {
            "A": 50,
            "B": 25,
            "C": 12,
            "D": 20,
            "E": 18,
            "F": 12,
            "G": 12,
            "H": 12,
            "I": 18,
            "J": 80,
            "K": 40,
            "L": 100,
            "M": 120,
            "N": 120,
            "O": 100,
            "P": 100,
            "Q": 16,
            "R": 12,
            "S": 14,
            "T": 12,
            "U": 25,
            "V": 14,
            "W": 150,
        }
        for column, width in column_widths.items():
            summary_sheet.column_dimensions[column].width = width

        for cell in summary_sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in summary_sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for cell in stats_sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        stats_sheet.column_dimensions["A"].width = 24
        stats_sheet.column_dimensions["B"].width = 24

    print(f"Excel 报告已生成: {OUTPUT_FILE}")
    print(f"共包含 {len(summaries)} 篇论文，输出 2 个工作表")
    return True


if __name__ == "__main__":
    if create_excel_for_policy_analysis():
        print("政策分析项目 Excel 文件生成完成")
    else:
        print("政策分析项目 Excel 文件生成失败")
